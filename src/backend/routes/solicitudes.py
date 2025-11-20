from __future__ import annotations

import json
from datetime import datetime
from typing import Any, Iterable
from io import BytesIO

from flask import Blueprint, jsonify, request, send_file

from ..core.db import get_connection
from ..models.schemas import BudgetIncreaseDecision, SolicitudCreate, SolicitudDraft
from ..services.auth.auth import authenticate_request, get_current_user_id
from ..models.roles import has_role

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle


bp = Blueprint("solicitudes", __name__, url_prefix="/api")

STATUS_PENDING = "pendiente_de_aprobacion"
STATUS_APPROVED = "aprobada"
STATUS_REJECTED = "rechazada"
STATUS_CANCELLED = "cancelada"
STATUS_FINALIZED = "finalizada"
STATUS_DRAFT = "draft"
STATUS_CANCEL_PENDING = "cancelacion_pendiente"
STATUS_CANCEL_REJECTED = "cancelacion_rechazada"
STATUS_IN_TREATMENT = "en_tratamiento"


def _utcnow_iso() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def _require_auth() -> str | None:
    uid = get_current_user_id()
    if uid:
        return uid
    authenticate_request()
    uid = get_current_user_id()
    return uid


def _json_error(code: str, message: str, status: int = 400):
    return jsonify({"ok": False, "error": {"code": code, "message": message}}), status


def _coerce_str(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _map_criticidad(value: Any) -> str | None:
    """Normaliza valores de criticidad del frontend a lo que espera el modelo Pydantic."""
    v = _coerce_str(value).lower()
    if not v:
        return None
    if v in {"alta", "high", "crítica", "critica"}:
        return "Alta"
    return "Normal"


def _get_payload_from_request() -> dict[str, Any]:
    """Unifica la lectura del payload: JSON o FormData."""
    content_type = (request.content_type or "").lower()

    if "application/json" in content_type:
        return request.get_json(force=True, silent=False) or {}

    # FormData case
    form = request.form.to_dict(flat=True)
    payload = {
        "centro": form.get("centro"),
        "sector": form.get("sector"),
        "justificacion": form.get("justificacion"),
        "almacen_virtual": form.get("almacen_virtual") or form.get("almacen"),
        "fecha_necesidad": form.get("fecha_necesidad") or form.get("fechaNecesaria") or form.get("fecha_necesaria"),
        "centro_costos": form.get("centro_costos") or form.get("centroCostos"),
        "criticidad": _map_criticidad(form.get("criticidad")),
    }

    # Handle items - FormData sends items as a JSON string
    items_str = form.get("items")
    if items_str:
        try:
            payload["items"] = json.loads(items_str)
        except (json.JSONDecodeError, TypeError):
            payload["items"] = []

    return payload


def _fetch_user(con, uid: str | None):
    if not uid:
        return None
    return con.execute(
        """
        SELECT id_spm, nombre, apellido, rol, centros, jefe, gerente1, gerente2
          FROM usuarios
         WHERE lower(id_spm)=?
        """,
        (uid.lower(),),
    ).fetchone()


def _normalize_uid(value: Any) -> str | None:
    normalized = _coerce_str(value).lower()
    return normalized or None


def _ensure_user_exists(con, uid: str | None) -> str | None:
    """Return a normalized user id only if it exists in usuarios."""
    normalized = _normalize_uid(uid)
    if not normalized:
        return None
    row = con.execute(
        "SELECT 1 FROM usuarios WHERE lower(id_spm)=?",
        (normalized,),
    ).fetchone()
    return normalized if row else None


def _validar_material_existe(con, codigo: str) -> bool:
    """Verificar si un código de material existe en la tabla materiales."""
    if not codigo or not isinstance(codigo, str):
        return False
    codigo = codigo.strip()
    if not codigo:
        return False
    row = con.execute(
        "SELECT 1 FROM materiales WHERE codigo = ? LIMIT 1",
        (codigo,),
    ).fetchone()
    return row is not None


def _resolve_approver(con, user: dict[str, Any] | None, total_monto: float = 0.0) -> str | None:
    if not user:
        return None
    
    # Determinar el aprobador basado en el monto total
    approver_field, _, _ = _get_approver_config(total_monto)
    
    approver_email = _coerce_str(user.get(approver_field))
    if approver_email:
        # Buscar el id_spm del usuario con este email
        approver_user = con.execute(
            "SELECT id_spm FROM usuarios WHERE lower(mail) = ?",
            (approver_email.lower(),)
        ).fetchone()
        if approver_user:
            approver_id = approver_user["id_spm"]
            # FIX #2: Validar que el aprobador existe y está activo
            if _ensure_approver_exists_and_active(con, approver_id):
                return approver_id
    
    # Fallback: buscar en otros campos si el campo específico no está disponible
    for field in ("jefe", "gerente1", "gerente2"):
        approver_email = _coerce_str(user.get(field))
        if approver_email:
            approver_user = con.execute(
                "SELECT id_spm FROM usuarios WHERE lower(mail) = ?",
                (approver_email.lower(),)
            ).fetchone()
            if approver_user:
                approver_id = approver_user["id_spm"]
                # FIX #2: Validar que el aprobador existe y está activo
                if _ensure_approver_exists_and_active(con, approver_id):
                    return approver_id
    return None


def _resolve_planner(user: dict[str, Any] | None, con=None) -> str | None:
    if not user:
        return None
    for field in ("gerente2", "gerente1"):
        value = _coerce_str(user.get(field))
        if value:
            planner_id = value.lower()
            # FIX #3: Validar que el planificador existe y está disponible
            if con is not None:
                if _ensure_planner_exists_and_available(con, planner_id):
                    return planner_id
            else:
                # Si no hay conexión, retornar el valor como estaba antes
                return planner_id
    return None


def _normalize_items(raw_items: Iterable[Any], con=None) -> tuple[list[dict[str, Any]], float]:
    items: list[dict[str, Any]] = []
    total = 0.0
    invalid_materials = []
    
    for raw in raw_items or []:
        if not isinstance(raw, dict):
            continue
        codigo = _coerce_str(raw.get("codigo"))
        if not codigo:
            continue
        
        # FIX #1: Validar que el material exista en el catálogo
        if con is not None and not _validar_material_existe(con, codigo):
            invalid_materials.append(codigo)
            continue
            
        descripcion = _coerce_str(raw.get("descripcion"))
        try:
            cantidad = int(raw.get("cantidad", 0))
        except (TypeError, ValueError):
            cantidad = 0
        if cantidad < 1:
            cantidad = 1
        precio_raw = raw.get("precio_unitario")
        if precio_raw is None:
            precio_raw = raw.get("precio")
        try:
            precio = float(precio_raw)
        except (TypeError, ValueError):
            precio = 0.0
        if precio < 0:
            precio = 0.0
        subtotal = round(cantidad * precio, 2)
        item: dict[str, Any] = {
            "codigo": codigo,
            "descripcion": descripcion,
            "cantidad": cantidad,
            "precio_unitario": round(precio, 2),
            "comentario": raw.get("comentario"),
            "subtotal": subtotal,
        }
        unidad = raw.get("unidad") or raw.get("uom") or raw.get("unidad_medida")
        if unidad:
            item["unidad"] = _coerce_str(unidad)
        items.append(item)
        total += subtotal
    
    # Si hay materiales inválidos y se validaba, lanzar error
    if invalid_materials and con is not None:
        raise ValueError(f"Los siguientes códigos de material no existen en el catálogo: {', '.join(invalid_materials)}")
    
    return items, round(total, 2)


def _parse_draft_payload(uid: str, payload: dict[str, Any]) -> dict[str, Any]:
    model = SolicitudDraft(**{**payload, "id_usuario": uid})
    data = model.model_dump()
    data["id_usuario"] = uid.lower()
    fecha = data.get("fecha_necesidad")
    if fecha:
        data["fecha_necesidad"] = fecha.isoformat()
    data["criticidad"] = data.get("criticidad") or "Normal"
    return data


def _parse_full_payload(uid: str, payload: dict[str, Any], *, expect_items: bool, con=None) -> dict[str, Any]:
    sanitized_items = []
    for item in payload.get("items", []) or []:
        if not isinstance(item, dict):
            continue
        sanitized_items.append(
            {
                "codigo": item.get("codigo"),
                "descripcion": item.get("descripcion"),
                "cantidad": item.get("cantidad"),
                "precio_unitario": item.get("precio_unitario"),
                "comentario": item.get("comentario"),
            }
        )
    payload_for_model = {**payload, "items": sanitized_items, "id_usuario": uid}
    model = SolicitudCreate(**payload_for_model)
    data = model.model_dump()
    data["id_usuario"] = uid.lower()
    fecha = data.get("fecha_necesidad")
    if fecha:
        data["fecha_necesidad"] = fecha.isoformat()
    data["criticidad"] = data.get("criticidad") or "Normal"
    
    # Usar la nueva versión de _normalize_items con validación
    items, total = _normalize_items(payload.get("items", []), con=con)
    if expect_items and not items:
        raise ValueError("Debe incluir al menos un ítem")
    data["items"] = items
    data["total_monto"] = total
    return data


def _json_load(raw: Any) -> dict[str, Any]:
    if isinstance(raw, dict):
        return dict(raw)
    if not raw:
        return {}
    try:
        return json.loads(raw)
    except Exception:
        return {}


def _serialize_items(items: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for raw in items or []:
        if not isinstance(raw, dict):
            continue
        cantidad = raw.get("cantidad")
        try:
            cantidad_int = int(cantidad)
        except (TypeError, ValueError):
            cantidad_int = 0
        if cantidad_int < 1:
            cantidad_int = 1
        try:
            precio = float(raw.get("precio_unitario", 0.0))
        except (TypeError, ValueError):
            precio = 0.0
        subtotal = raw.get("subtotal")
        if subtotal is None:
            subtotal = round(cantidad_int * precio, 2)
        item = {
            "codigo": _coerce_str(raw.get("codigo")),
            "descripcion": _coerce_str(raw.get("descripcion")),
            "cantidad": cantidad_int,
            "precio_unitario": round(precio, 2),
            "comentario": raw.get("comentario"),
            "subtotal": round(float(subtotal), 2),
        }
        unidad = raw.get("unidad") or raw.get("uom")
        if unidad:
            item["unidad"] = _coerce_str(unidad)
        result.append(item)
    return result


def _ensure_totals(data: dict[str, Any], fallback: float) -> float:
    try:
        stored = float(data.get("total_monto", fallback))
    except (TypeError, ValueError):
        stored = fallback
    if stored <= 0 and data.get("items"):
        subtotal = sum(item.get("subtotal", 0.0) for item in data.get("items", []))
        try:
            stored = float(subtotal)
        except (TypeError, ValueError):
            stored = fallback
    data["total_monto"] = round(stored, 2)
    return data["total_monto"]


def _assign_planner_automatically(con, centro: str, sector: str, almacen_virtual: str) -> str | None:
    """Asigna automáticamente un planificador basado en Centro, Sector y Almacén Virtual."""
    if not centro or not sector or not almacen_virtual:
        return None
    
    # Buscar asignación específica por centro, sector y almacen
    row = con.execute(
        """
        SELECT p.usuario_id
          FROM planificador_asignaciones pa
          JOIN planificadores p ON pa.planificador_id = p.usuario_id
         WHERE pa.centro = ? AND pa.sector = ? AND pa.almacen_virtual = ?
         ORDER BY pa.created_at ASC
         LIMIT 1
        """,
        (centro, sector, almacen_virtual),
    ).fetchone()
    
    if row:
        return row["usuario_id"]
    
    # Si no hay asignación específica, buscar por centro y sector
    row = con.execute(
        """
        SELECT p.usuario_id
          FROM planificador_asignaciones pa
          JOIN planificadores p ON pa.planificador_id = p.usuario_id
         WHERE pa.centro = ? AND pa.sector = ? AND pa.almacen_virtual IS NULL
         ORDER BY pa.created_at ASC
         LIMIT 1
        """,
        (centro, sector),
    ).fetchone()
    
    if row:
        return row["usuario_id"]
    
    # Si no hay asignación por centro y sector, buscar solo por centro
    row = con.execute(
        """
        SELECT p.usuario_id
          FROM planificador_asignaciones pa
          JOIN planificadores p ON pa.planificador_id = p.usuario_id
         WHERE pa.centro = ? AND pa.sector IS NULL AND pa.almacen_virtual IS NULL
         ORDER BY pa.created_at ASC
         LIMIT 1
        """,
        (centro,),
    ).fetchone()
    
    return row["usuario_id"] if row else None


def _get_approver_config(total_monto: float = 0.0) -> tuple[str, float, float]:
    """Determinar el nivel de aprobación requerido basado en el monto.
    
    Retorna: (approver_field, min_monto, max_monto)
    - jefe: USD 0.01 a 20000
    - gerente1: USD 20000.01 a 100000
    - gerente2: USD 100000.01 en adelante
    """
    if total_monto <= 0:
        return ("jefe", 0.0, 20000.0)
    elif total_monto <= 20000.0:
        return ("jefe", 0.0, 20000.0)
    elif total_monto <= 100000.0:
        return ("gerente1", 20000.01, 100000.0)
    else:
        return ("gerente2", 100000.01, float("inf"))


def _ensure_approver_exists_and_active(con, approver_id: str | None) -> bool:
    """Verificar que el aprobador existe y está activo en el sistema.
    
    FIX #2: Valida que no haya aprobadores fantasma o inactivos.
    """
    if not approver_id:
        return False
    
    approver_id = _coerce_str(approver_id).lower().strip()
    if not approver_id:
        return False
    
    row = con.execute(
        """
        SELECT id_spm, estado_registro 
        FROM usuarios 
        WHERE lower(id_spm) = ? OR lower(mail) = ?
        LIMIT 1
        """,
        (approver_id, approver_id)
    ).fetchone()
    
    if not row:
        return False
    
    # Convertir a dict si es necesario
    if hasattr(row, 'get'):
        estado = _coerce_str(row.get("estado_registro", "")).lower()
    else:
        # Acceso por índice si es tupla
        estado = _coerce_str(row[1] if len(row) > 1 else "").lower()
    
    # Verificar que el usuario esté activo (estado_registro puede ser 'Activo', 'activo', 'A', etc.)
    if estado and estado not in ("activo", "active", "a", "1", "true"):
        return False
    
    return True


def _ensure_planner_exists_and_available(con, planner_id: str | None) -> bool:
    """Verificar que el planificador existe, está activo y disponible en el sistema.
    
    FIX #3: Valida que no haya planificadores fantasma, inactivos o sobrecargados.
    """
    if not planner_id:
        return False
    
    planner_id = _coerce_str(planner_id).lower().strip()
    if not planner_id:
        return False
    
    row = con.execute(
        """
        SELECT id_spm, estado_registro, rol
        FROM usuarios 
        WHERE lower(id_spm) = ? OR lower(mail) = ?
        LIMIT 1
        """,
        (planner_id, planner_id)
    ).fetchone()
    
    if not row:
        return False
    
    # Convertir a dict si es necesario
    if hasattr(row, 'get'):
        estado = _coerce_str(row.get("estado_registro", "")).lower()
        rol = _coerce_str(row.get("rol", "")).lower()
    else:
        # Acceso por índice si es tupla
        estado = _coerce_str(row[1] if len(row) > 1 else "").lower()
        rol = _coerce_str(row[2] if len(row) > 2 else "").lower()
    
    # Verificar que el usuario esté activo
    if estado and estado not in ("activo", "active", "a", "1", "true"):
        return False
    
    # Verificar que tenga rol de planificador
    if rol and rol not in ("planificador", "planner", "gerente", "gerente1", "gerente2", "admin", "administrador"):
        return False
    
    # Verificar que no esté sobrecargado (máximo N solicitudes en tratamiento por planificador)
    MAX_ACTIVE_SOLICITUDES = 20
    count_result = con.execute(
        """
        SELECT COUNT(*) as count
        FROM solicitudes
        WHERE lower(planner_id) = ? AND status IN (?, ?)
        """,
        (planner_id, STATUS_IN_TREATMENT, STATUS_APPROVED)
    ).fetchone()
    
    if hasattr(count_result, 'get'):
        active_count = count_result.get("count", 0) if count_result else 0
    else:
        active_count = count_result[0] if count_result else 0
    
    if active_count >= MAX_ACTIVE_SOLICITUDES:
        return False
    
    return True


def _pre_validar_aprobacion(con, row: dict[str, Any], approver_user: dict[str, Any] | None) -> tuple[bool, str | None]:
    """Validaciones previas a la aprobación de una solicitud.
    
    FIX #4: Realiza 5 validaciones críticas antes de permitir aprobación.
    
    Retorna: (es_valido, mensaje_error)
    """
    if not row:
        return False, "Solicitud no encontrada"
    
    sol_id = row.get("id")
    data = _json_load(row.get("data_json", "{}"))
    items = data.get("items", [])
    total_monto = row.get("total_monto", 0.0)
    
    # Validación 1: Aprobador activo
    if approver_user:
        approver_id = approver_user.get("id_spm")
        if not _ensure_approver_exists_and_active(con, approver_id):
            return False, f"El aprobador {approver_id} no está activo en el sistema"
    
    # Validación 2: Materiales válidos (todos en el catálogo)
    invalid_materials = []
    for item in items:
        codigo = _coerce_str(item.get("codigo", "")).strip()
        if codigo and not _validar_material_existe(con, codigo):
            invalid_materials.append(codigo)
    if invalid_materials:
        return False, f"Materiales inválidos: {', '.join(invalid_materials[:5])}"
    
    # Validación 3: Total consistente (no debe ser 0 o negativo)
    if not isinstance(total_monto, (int, float)) or total_monto <= 0:
        return False, f"Monto total inválido: {total_monto}"
    
    # Validación 4: Presupuesto disponible (sin exceder límites del aprobador)
    if approver_user:
        _, min_monto, max_monto = _get_approver_config(total_monto)
        if total_monto < min_monto or total_monto > max_monto:
            return False, f"Monto USD ${total_monto:,.2f} fuera del rango autorizado por este aprobador (USD ${min_monto:,.2f} - USD ${max_monto:,.2f})"
    
    # Validación 5: Usuario solicitante activo
    usuario_id = row.get("id_usuario")
    if usuario_id:
        usuario_row = con.execute(
            "SELECT estado_registro FROM usuarios WHERE lower(id_spm) = ? OR lower(mail) = ?",
            (usuario_id.lower(), usuario_id.lower())
        ).fetchone()
        if usuario_row:
            # Manejar tanto dict como tuplas
            if hasattr(usuario_row, 'get'):
                estado = _coerce_str(usuario_row.get("estado_registro", "")).lower()
            else:
                estado = _coerce_str(usuario_row[0] if len(usuario_row) > 0 else "").lower()
            
            if estado and estado not in ("activo", "active", "a", "1", "true"):
                return False, f"Usuario solicitante {usuario_id} no está activo"
        else:
            return False, f"Usuario solicitante {usuario_id} no encontrado"
    
    return True, None


def _serialize_row(row: dict[str, Any], *, detailed: bool) -> dict[str, Any]:
    data = _json_load(row.get("data_json"))
    data.setdefault("items", [])
    items = _serialize_items(data.get("items"))
    data["items"] = items
    total = _ensure_totals(data, float(row.get("total_monto") or 0.0))
    base: dict[str, Any] = {
        "id": row.get("id"),
        "status": row.get("status"),
        "centro": row.get("centro"),
        "sector": row.get("sector"),
        "justificacion": row.get("justificacion"),
        "centro_costos": row.get("centro_costos"),
        "almacen_virtual": row.get("almacen_virtual"),
        "criticidad": row.get("criticidad") or data.get("criticidad"),
        "fecha_necesidad": row.get("fecha_necesidad") or data.get("fecha_necesidad"),
        "id_usuario": row.get("id_usuario"),
        "aprobador_id": row.get("aprobador_id") or data.get("aprobador_id"),
        "planner_id": row.get("planner_id") or data.get("planner_id"),
        "total_monto": total,
        "created_at": row.get("created_at"),
        "updated_at": row.get("updated_at"),
        "notificado_at": row.get("notificado_at"),
        "data_json": data,
    }
    cancel_reason = data.get("cancel_reason")
    if cancel_reason:
        base["cancel_reason"] = cancel_reason
    if data.get("cancelled_at"):
        base["cancelled_at"] = data.get("cancelled_at")
    cancel_request = data.get("cancel_request")
    if isinstance(cancel_request, dict):
        base["cancel_request"] = cancel_request
    if detailed:
        base["items"] = items
    return base


def _load_solicitud(con, sol_id: int):
    return con.execute(
        """
        SELECT id, id_usuario, centro, sector, justificacion, centro_costos, almacen_virtual,
               data_json, status, aprobador_id, total_monto, notificado_at,
               created_at, updated_at, criticidad, fecha_necesidad, planner_id
          FROM solicitudes
         WHERE id=?
        """,
        (sol_id,),
    ).fetchone()


def _create_notification(con, destinatario: str | None, solicitud_id: int, mensaje: str) -> None:
    dest = _coerce_str(destinatario).lower()
    if not dest:
        return
    con.execute(
        "INSERT INTO notificaciones (destinatario_id, solicitud_id, mensaje, leido) VALUES (?,?,?,0)",
        (dest, solicitud_id, mensaje),
    )


def _can_view(user: dict[str, Any] | None, row: dict[str, Any]) -> bool:
    uid = _coerce_str(user.get("id_spm")) if user else ""
    if uid and uid.lower() == _coerce_str(row.get("id_usuario")).lower():
        return True
    approver = _coerce_str(row.get("aprobador_id")).lower()
    if uid and approver and uid.lower() == approver:
        return True
    planner = _coerce_str(row.get("planner_id")).lower()
    if uid and planner and uid.lower() == planner:
        return True
    if has_role(user, "planner", "planificador", "admin", "administrador"):
        return True
    return False


def _can_decide_cancel(user: dict[str, Any] | None, row: dict[str, Any]) -> bool:
    if not user:
        return False
    uid = _coerce_str(user.get("id_spm")).lower()
    if not uid:
        return False
    approver = _coerce_str(row.get("aprobador_id")).lower()
    planner = _coerce_str(row.get("planner_id")).lower()
    if uid == approver or uid == planner:
        return True
    return has_role(user, "planner", "planificador", "admin", "administrador")


def _can_resolve(user: dict[str, Any] | None, row: dict[str, Any]) -> bool:
    if not user:
        return False
    uid = _coerce_str(user.get("id_spm")).lower()
    if not uid:
        return False
    approver = _coerce_str(row.get("aprobador_id")).lower()
    planner = _coerce_str(row.get("planner_id")).lower()
    if uid == approver or uid == planner:
        return True
    return has_role(user, "planner", "planificador", "admin", "administrador", "aprobador")


@bp.get("/solicitudes")
def listar_solicitudes():
    uid = _require_auth()
    if not uid:
        return _json_error("NOAUTH", "No autenticado", 401)
    with get_connection() as con:
        con.row_factory = lambda cursor, row: {col[0]: row[idx] for idx, col in enumerate(cursor.description)}
        rows = con.execute(
            """
            SELECT id, id_usuario, centro, sector, justificacion, centro_costos, almacen_virtual,
                   data_json, status, aprobador_id, total_monto, notificado_at,
                   created_at, updated_at, criticidad, fecha_necesidad, planner_id
              FROM solicitudes
             WHERE lower(id_usuario)=?
          ORDER BY datetime(created_at) DESC, id DESC
            """,
            (uid.lower(),),
        ).fetchall()
    items = [_serialize_row(row, detailed=False) for row in rows]
    return {"ok": True, "items": items, "total": len(items)}


@bp.get("/solicitudes/<int:sol_id>")
def obtener_solicitud(sol_id: int):
    uid = _require_auth()
    if not uid:
        return _json_error("NOAUTH", "No autenticado", 401)
    with get_connection() as con:
        con.row_factory = lambda cursor, row: {col[0]: row[idx] for idx, col in enumerate(cursor.description)}
        row = _load_solicitud(con, sol_id)
        if not row:
            return _json_error("NOTFOUND", "Solicitud no encontrada", 404)
        user = _fetch_user(con, uid)
        if not _can_view(user, row):
            return _json_error("FORBIDDEN", "No tienes acceso a esta solicitud", 403)
        solicitud = _serialize_row(row, detailed=True)
        # Agregar nombre del aprobador si existe
        aprobador_id = solicitud.get("aprobador_id")
        if aprobador_id:
            aprobador_user = _fetch_user(con, aprobador_id)
            if aprobador_user:
                solicitud["aprobador_nombre"] = f"{aprobador_user['nombre']} {aprobador_user['apellido']}"
        
        # Agregar nombre del planificador asignado si existe
        planner_id = solicitud.get("planner_id")
        if planner_id:
            planner_user = _fetch_user(con, planner_id)
            if planner_user:
                solicitud["planner_nombre"] = f"{planner_user['nombre']} {planner_user['apellido']}"
    return {"ok": True, "solicitud": solicitud}


def _sync_columns_from_payload(payload: dict[str, Any]) -> tuple[str, str, str, str, str, str, str]:
    return (
        payload.get("centro"),
        payload.get("sector"),
        payload.get("justificacion"),
        payload.get("centro_costos"),
        payload.get("almacen_virtual"),
        payload.get("criticidad") or "Normal",
        payload.get("fecha_necesidad"),
    )


@bp.route("/solicitudes/drafts", methods=["POST", "OPTIONS"])
def crear_borrador():
    if request.method == "OPTIONS":
        return "", 204
    uid = _require_auth()
    if not uid:
        return _json_error("NOAUTH", "No autenticado", 401)
    # NUEVO: unificamos la lectura del cuerpo
    payload = _get_payload_from_request()
    try:
        draft_data = _parse_draft_payload(uid, payload)
    except Exception as exc:  # validation error
        return _json_error("BAD_REQUEST", str(exc), 400)
    with get_connection() as con:
        con.row_factory = lambda cursor, row: {col[0]: row[idx] for idx, col in enumerate(cursor.description)}
        user = _fetch_user(con, uid)
        if not user:
            return _json_error("NOUSER", "Usuario no encontrado", 404)
        approver = _ensure_user_exists(con, _resolve_approver(con, user, 0.0))
        planner = _ensure_user_exists(con, _resolve_planner(user, con=con))
        draft_payload = {
            **draft_data,
            "items": [],
            "total_monto": 0.0,
        }
        if approver:
            draft_payload["aprobador_id"] = approver
        elif "aprobador_id" in draft_payload:
            draft_payload.pop("aprobador_id", None)
        if planner:
            draft_payload["planner_id"] = planner
        elif "planner_id" in draft_payload:
            draft_payload.pop("planner_id", None)
        data_json = json.dumps(draft_payload, ensure_ascii=False)
        centro, sector, justificacion, centro_costos, almacen_virtual, criticidad, fecha_necesidad = _sync_columns_from_payload(draft_payload)
        try:
            cur = con.execute(
                """
                INSERT INTO solicitudes (
                    id_usuario, centro, sector, justificacion, centro_costos, almacen_virtual,
                    data_json, status, aprobador_id, total_monto, criticidad, fecha_necesidad, planner_id
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    uid.lower(),
                    centro,
                    sector,
                    justificacion,
                    centro_costos,
                    almacen_virtual,
                    data_json,
                    STATUS_DRAFT,
                    approver,
                    0.0,
                    criticidad,
                    fecha_necesidad,
                    planner,
                ),
            )
            sol_id = cur.lastrowid
            con.commit()
        except Exception as exc:
            con.rollback()
            return _json_error("DB_ERROR", f"No se pudo crear el borrador: {exc}", 500)
    return {"ok": True, "id": sol_id, "solicitud_id": sol_id, "status": STATUS_DRAFT}


@bp.route("/solicitudes/<int:sol_id>/draft", methods=["PATCH", "OPTIONS"])
def actualizar_borrador(sol_id: int):
    if request.method == "OPTIONS":
        return "", 204
    uid = _require_auth()
    if not uid:
        return _json_error("NOAUTH", "No autenticado", 401)
    # NUEVO: unificamos la lectura del cuerpo
    payload = _get_payload_from_request()
    try:
        draft_data = _parse_full_payload(uid, payload, expect_items=False)
    except ValueError as exc:
        return _json_error("BAD_REQUEST", str(exc), 400)
    except Exception as exc:
        return _json_error("BAD_REQUEST", str(exc), 400)
    with get_connection() as con:
        con.row_factory = lambda cursor, row: {col[0]: row[idx] for idx, col in enumerate(cursor.description)}
        row = _load_solicitud(con, sol_id)
        if not row:
            return _json_error("NOTFOUND", "Solicitud no encontrada", 404)
        if _coerce_str(row.get("id_usuario")).lower() != uid.lower():
            return _json_error("FORBIDDEN", "No puedes editar este borrador", 403)
        if row.get("status") not in (STATUS_DRAFT, STATUS_CANCEL_REJECTED):
            return _json_error("INVALID_STATE", "La solicitud no está en borrador", 409)
        existing_data = _json_load(row.get("data_json"))
        existing_data.update({k: v for k, v in draft_data.items() if k != "items"})
        if draft_data.get("items"):
            existing_data["items"] = draft_data["items"]
            existing_data["total_monto"] = draft_data["total_monto"]
        
        # Recalcular aprobador si cambió el monto
        new_total = existing_data.get("total_monto", 0.0)
        old_total = row.get("total_monto", 0.0)
        if abs(new_total - old_total) > 0.01:  # Pequeña tolerancia para flotantes
            user = _fetch_user(con, uid)
            new_approver = _ensure_user_exists(con, _resolve_approver(con, user, new_total))
            if new_approver != row.get("aprobador_id"):
                existing_data["aprobador_id"] = new_approver
        
        centro, sector, justificacion, centro_costos, almacen_virtual, criticidad, fecha_necesidad = _sync_columns_from_payload(draft_data)
        data_json = json.dumps(existing_data, ensure_ascii=False)
        try:
            con.execute(
                """
                UPDATE solicitudes
                   SET centro=?, sector=?, justificacion=?, centro_costos=?, almacen_virtual=?,
                       data_json=?, total_monto=?, aprobador_id=?, criticidad=?, fecha_necesidad=?,
                       updated_at=CURRENT_TIMESTAMP
                 WHERE id=?
                """,
                (
                    centro,
                    sector,
                    justificacion,
                    centro_costos,
                    almacen_virtual,
                    data_json,
                    existing_data.get("total_monto", row.get("total_monto")),
                    existing_data.get("aprobador_id", row.get("aprobador_id")),
                    criticidad,
                    fecha_necesidad,
                    sol_id,
                ),
            )
            con.commit()
        except Exception as exc:
            con.rollback()
            return _json_error("DB_ERROR", f"No se pudo guardar el borrador: {exc}", 500)
    return {"ok": True, "id": sol_id, "status": row.get("status")}


def _finalizar_solicitud(con, row: dict[str, Any], final_data: dict[str, Any], user: dict[str, Any] | None, *, is_new: bool) -> tuple[int, dict[str, Any]]:
    approver = _ensure_user_exists(
        con,
        final_data.get("aprobador_id") or _resolve_approver(con, user, final_data.get("total_monto", 0.0)),
    )
    planner = _ensure_user_exists(
        con,
        final_data.get("planner_id") or row.get("planner_id") or _resolve_planner(user, con=con),
    )
    final_payload = {**_json_load(row.get("data_json")), **final_data}
    final_payload["aprobador_id"] = approver
    if planner:
        final_payload["planner_id"] = planner
    elif "planner_id" in final_payload:
        final_payload.pop("planner_id", None)
    final_payload.pop("cancel_request", None)
    final_payload.pop("cancel_reason", None)
    final_payload.pop("cancelled_at", None)
    final_payload["total_monto"] = final_data["total_monto"]
    data_json = json.dumps(final_payload, ensure_ascii=False)
    centro, sector, justificacion, centro_costos, almacen_virtual, criticidad, fecha_necesidad = _sync_columns_from_payload(final_payload)
    now_iso = _utcnow_iso()
    if is_new:
        cur = con.execute(
            """
            INSERT INTO solicitudes (
                id_usuario, centro, sector, justificacion, centro_costos, almacen_virtual,
                data_json, status, aprobador_id, total_monto, criticidad, fecha_necesidad, planner_id, notificado_at
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                final_payload["id_usuario"],
                centro,
                sector,
                justificacion,
                centro_costos,
                almacen_virtual,
                data_json,
                STATUS_PENDING,
                approver,
                final_payload["total_monto"],
                criticidad,
                fecha_necesidad,
                planner,
                now_iso,
            ),
        )
        sol_id = cur.lastrowid
    else:
        con.execute(
            """
            UPDATE solicitudes
               SET centro=?, sector=?, justificacion=?, centro_costos=?, almacen_virtual=?,
                   data_json=?, status=?, aprobador_id=?, total_monto=?, criticidad=?,
                   fecha_necesidad=?, planner_id=?, notificado_at=?, updated_at=CURRENT_TIMESTAMP
             WHERE id=?
            """,
            (
                centro,
                sector,
                justificacion,
                centro_costos,
                almacen_virtual,
                data_json,
                STATUS_PENDING,
                approver,
                final_payload["total_monto"],
                criticidad,
                fecha_necesidad,
                planner,
                now_iso,
                row["id"],
            ),
        )
        sol_id = row["id"]
    if approver:
        _create_notification(con, approver, sol_id, f"Solicitud #{sol_id} pendiente de aprobación")
    return sol_id, final_payload


@bp.route("/solicitudes/<int:sol_id>", methods=["PUT", "OPTIONS"])
def finalizar_solicitud(sol_id: int):
    if request.method == "OPTIONS":
        return "", 204
    uid = _require_auth()
    if not uid:
        return _json_error("NOAUTH", "No autenticado", 401)
    payload = request.get_json(force=True, silent=False) or {}
    
    with get_connection() as con:
        con.row_factory = lambda cursor, row: {col[0]: row[idx] for idx, col in enumerate(cursor.description)}
        try:
            # Pasar la conexión para validar materiales
            final_data = _parse_full_payload(uid, payload, expect_items=True, con=con)
        except ValueError as exc:
            return _json_error("BAD_REQUEST", str(exc), 400)
        except Exception as exc:
            return _json_error("BAD_REQUEST", str(exc), 400)
        
        row = _load_solicitud(con, sol_id)
        if not row:
            return _json_error("NOTFOUND", "Solicitud no encontrada", 404)
        if _coerce_str(row.get("id_usuario")).lower() != uid.lower():
            return _json_error("FORBIDDEN", "No puedes finalizar esta solicitud", 403)
        if row.get("status") not in (STATUS_DRAFT, STATUS_CANCEL_REJECTED):
            return _json_error("INVALID_STATE", "La solicitud no está en borrador", 409)
        user = _fetch_user(con, uid)
        try:
            sol_id, final_payload = _finalizar_solicitud(con, row, final_data, user, is_new=False)
            con.commit()
        except Exception as exc:
            con.rollback()
            return _json_error("DB_ERROR", f"No se pudo enviar la solicitud: {exc}", 500)
    return {"ok": True, "id": sol_id, "status": STATUS_PENDING, "total_monto": final_payload.get("total_monto")}


@bp.route("/solicitudes", methods=["POST", "OPTIONS"])
def crear_solicitud():
    if request.method == "OPTIONS":
        return "", 204
    uid = _require_auth()
    if not uid:
        return _json_error("NOAUTH", "No autenticado", 401)
    # NUEVO: unificamos la lectura del cuerpo
    payload = _get_payload_from_request()

    with get_connection() as con:
        con.row_factory = lambda cursor, row: {col[0]: row[idx] for idx, col in enumerate(cursor.description)}
        try:
            # Pasar la conexión para validar materiales
            final_data = _parse_full_payload(uid, payload, expect_items=True, con=con)
        except ValueError as exc:
            return _json_error("BAD_REQUEST", str(exc), 400)
        except Exception as exc:
            return _json_error("BAD_REQUEST", str(exc), 400)
        
        dummy_row = {
            "id": None,
            "data_json": json.dumps({}, ensure_ascii=False),
            "planner_id": None,
        }
        final_data["id_usuario"] = uid.lower()
        user = _fetch_user(con, uid)
        try:
            sol_id, final_payload = _finalizar_solicitud(con, dummy_row, final_data, user, is_new=True)
            con.commit()
        except Exception as exc:
            con.rollback()
            return _json_error("DB_ERROR", f"No se pudo crear la solicitud: {exc}", 500)
    return {"ok": True, "id": sol_id, "status": STATUS_PENDING, "total_monto": final_payload.get("total_monto")}


def _handle_direct_cancel(con, row: dict[str, Any], reason: str | None) -> dict[str, Any]:
    data = _json_load(row.get("data_json"))
    data["cancel_reason"] = reason or data.get("cancel_reason")
    data["cancelled_at"] = _utcnow_iso()
    data.pop("cancel_request", None)
    data_json = json.dumps(data, ensure_ascii=False)
    con.execute(
        """
        UPDATE solicitudes
           SET status=?, data_json=?, updated_at=CURRENT_TIMESTAMP
         WHERE id=?
        """,
        (STATUS_CANCELLED, data_json, row["id"]),
    )
    return data


def _create_cancel_request(row: dict[str, Any], reason: str | None, uid: str) -> dict[str, Any]:
    cancel_request = {
        "status": "pendiente",
        "reason": reason,
        "requested_at": _utcnow_iso(),
        "requested_by": uid.lower(),
    }
    return cancel_request


@bp.route("/solicitudes/<int:sol_id>/decidir", methods=["POST", "OPTIONS"])
def decidir_solicitud(sol_id: int):
    if request.method == "OPTIONS":
        return "", 204
    uid = _require_auth()
    if not uid:
        return _json_error("NOAUTH", "No autenticado", 401)
    payload = request.get_json(silent=True) or {}
    try:
        decision = BudgetIncreaseDecision(**payload)
    except Exception as exc:
        return _json_error("BAD_REQUEST", str(exc), 400)

    accion = decision.accion
    comentario = decision.comentario or None

    with get_connection() as con:
        con.row_factory = lambda cursor, row: {col[0]: row[idx] for idx, col in enumerate(cursor.description)}
        row = _load_solicitud(con, sol_id)
        if not row:
            return _json_error("NOTFOUND", "Solicitud no encontrada", 404)
        user = _fetch_user(con, uid)
        if not _can_resolve(user, row):
            return _json_error("FORBIDDEN", "No tienes permisos para esta operación", 403)
        
        # FIX #2: Validar que el usuario que está decidiendo existe y está activo
        if not _ensure_approver_exists_and_active(con, uid):
            return _json_error("INVALID_APPROVER", "Tu perfil de aprobador no está activo en el sistema", 403)
        
        if row.get("status") != STATUS_PENDING:
            return _json_error("INVALID_STATE", "La solicitud no está pendiente de aprobación", 409)
        
        # FIX #4: Pre-validaciones antes de permitir aprobación
        if decision.accion == "aprobar":
            es_valido, error_msg = _pre_validar_aprobacion(con, row, user)
            if not es_valido:
                return _json_error("VALIDATION_ERROR", error_msg or "Validación previa falló", 400)

        decision_at = _utcnow_iso()
        data = _json_load(row.get("data_json"))
        decision_payload = {
            "status": STATUS_APPROVED if accion == "aprobar" else STATUS_REJECTED,
            "accion": accion,
            "decided_at": decision_at,
            "decided_by": uid.lower(),
        }
        if comentario:
            decision_payload["comment"] = comentario
            data["decision_comment"] = comentario
        data["decision"] = decision_payload
        data.pop("cancel_request", None)

        # Determinar status final y asignar planificador si se aprueba
        assigned_planner_id = None
        if accion == "aprobar":
            # Asignar planificador automáticamente
            centro = row.get("centro")
            sector = row.get("sector") 
            almacen_virtual = row.get("almacen_virtual")
            assigned_planner_id = _assign_planner_automatically(con, centro, sector, almacen_virtual)
            
            if assigned_planner_id:
                status_final = STATUS_IN_TREATMENT
                data["assigned_planner"] = assigned_planner_id
                message = f"Solicitud #{sol_id} aprobada y asignada al planificador"
            else:
                status_final = STATUS_APPROVED
                message = f"Solicitud #{sol_id} aprobada (sin planificador asignado)"
        else:
            status_final = STATUS_REJECTED
            message = f"Solicitud #{sol_id} rechazada"

        try:
            data_json = json.dumps(data, ensure_ascii=False)
            con.execute(
                """
                UPDATE solicitudes
                   SET status=?, data_json=?, updated_at=CURRENT_TIMESTAMP,
                       notificado_at=?, aprobador_id=?, planner_id=?
                 WHERE id=?
                """,
                (status_final, data_json, decision_at, uid.lower(), 
                 assigned_planner_id, sol_id),
            )
            owner = row.get("id_usuario")
            planner = assigned_planner_id  # Usar el planificador asignado, no el anterior
            assigned_planner = data.get("assigned_planner")
            recipients = {owner, planner, assigned_planner}
            for dest in recipients:
                if dest:
                    _create_notification(con, dest, sol_id, message)
            con.commit()
        except Exception as exc:
            con.rollback()
            return _json_error("DB_ERROR", f"No se pudo registrar la decisión: {exc}", 500)

    return {"ok": True, "status": status_final, "decision": decision_payload}


@bp.route("/solicitudes/<int:sol_id>/cancel", methods=["PATCH", "OPTIONS"])
def cancelar_solicitud(sol_id: int):
    if request.method == "OPTIONS":
        return "", 204
    uid = _require_auth()
    if not uid:
        return _json_error("NOAUTH", "No autenticado", 401)
    payload = request.get_json(silent=True) or {}
    reason = _coerce_str(payload.get("reason")) or None
    with get_connection() as con:
        con.row_factory = lambda cursor, row: {col[0]: row[idx] for idx, col in enumerate(cursor.description)}
        row = _load_solicitud(con, sol_id)
        if not row:
            return _json_error("NOTFOUND", "Solicitud no encontrada", 404)
        owner = _coerce_str(row.get("id_usuario")).lower()
        if owner != uid.lower():
            return _json_error("FORBIDDEN", "No puedes cancelar esta solicitud", 403)
        status = row.get("status")
        try:
            if status in (STATUS_DRAFT, STATUS_CANCEL_REJECTED):
                data = _handle_direct_cancel(con, row, reason)
                con.commit()
                return {"ok": True, "status": STATUS_CANCELLED, "cancel_reason": data.get("cancel_reason")}
            if status == STATUS_CANCELLED:
                return _json_error("INVALID_STATE", "La solicitud ya está cancelada", 409)
            data = _json_load(row.get("data_json"))
            cancel_request = _create_cancel_request(row, reason, uid)
            data["cancel_request"] = cancel_request
            if reason:
                data["cancel_reason"] = reason
            data_json = json.dumps(data, ensure_ascii=False)
            con.execute(
                """
                UPDATE solicitudes
                   SET status=?, data_json=?, updated_at=CURRENT_TIMESTAMP
                 WHERE id=?
                """,
                (STATUS_CANCEL_PENDING, data_json, sol_id),
            )
            approver = row.get("aprobador_id") or row.get("planner_id")
            _create_notification(con, approver, sol_id, f"Solicitud #{sol_id} solicita cancelación")
            con.commit()
        except Exception as exc:
            con.rollback()
            return _json_error("DB_ERROR", f"No se pudo cancelar la solicitud: {exc}", 500)
    return {"ok": True, "status": STATUS_CANCEL_PENDING}


@bp.route("/solicitudes/<int:sol_id>/decidir_cancelacion", methods=["POST", "OPTIONS"])
def decidir_cancelacion(sol_id: int):
    if request.method == "OPTIONS":
        return "", 204
    uid = _require_auth()
    if not uid:
        return _json_error("NOAUTH", "No autenticado", 401)
    payload = request.get_json(force=True, silent=False) or {}
    accion = _coerce_str(payload.get("accion")).lower()
    comentario = _coerce_str(payload.get("comentario")) or None
    if accion not in {"aprobar", "rechazar"}:
        return _json_error("BAD_REQUEST", "Acción inválida", 400)
    with get_connection() as con:
        con.row_factory = lambda cursor, row: {col[0]: row[idx] for idx, col in enumerate(cursor.description)}
        row = _load_solicitud(con, sol_id)
        if not row:
            return _json_error("NOTFOUND", "Solicitud no encontrada", 404)
        user = _fetch_user(con, uid)
        if not _can_decide_cancel(user, row):
            return _json_error("FORBIDDEN", "No tienes permisos para esta operación", 403)
        if row.get("status") != STATUS_CANCEL_PENDING:
            return _json_error("INVALID_STATE", "La solicitud no está en cancelación pendiente", 409)
        data = _json_load(row.get("data_json"))
        cancel_request = data.get("cancel_request")
        if not isinstance(cancel_request, dict):
            cancel_request = {}
        cancel_request["decision_at"] = _utcnow_iso()
        cancel_request["decision_by"] = uid.lower()
        if comentario:
            cancel_request["decision_comment"] = comentario
        owner = row.get("id_usuario")
        try:
            if accion == "aprobar":
                cancel_request["status"] = "aprobada"
                data["cancel_request"] = cancel_request
                data["cancelled_at"] = cancel_request["decision_at"]
                data_json = json.dumps(data, ensure_ascii=False)
                con.execute(
                    """
                    UPDATE solicitudes
                       SET status=?, data_json=?, updated_at=CURRENT_TIMESTAMP
                     WHERE id=?
                    """,
                    (STATUS_CANCELLED, data_json, sol_id),
                )
                _create_notification(con, owner, sol_id, f"Solicitud #{sol_id} cancelada")
                result_status = STATUS_CANCELLED
            else:
                cancel_request["status"] = "rechazada"
                data["cancel_request"] = cancel_request
                data_json = json.dumps(data, ensure_ascii=False)
                con.execute(
                    """
                    UPDATE solicitudes
                       SET status=?, data_json=?, updated_at=CURRENT_TIMESTAMP
                     WHERE id=?
                    """,
                    (STATUS_CANCEL_REJECTED, data_json, sol_id),
                )
                _create_notification(con, owner, sol_id, f"Solicitud #{sol_id}: cancelación rechazada")
                result_status = STATUS_CANCEL_REJECTED
            con.commit()
        except Exception as exc:
            con.rollback()
            return _json_error("DB_ERROR", f"No se pudo registrar la decisión: {exc}", 500)
    return {"ok": True, "status": result_status, "accion": accion}


@bp.get("/solicitudes/export/excel")
def export_solicitudes_excel():
    """Exportar todas las solicitudes del usuario autenticado a Excel"""
    user_id = _require_auth()
    if not user_id:
        return _json_error("UNAUTHORIZED", "Autenticación requerida", 401)

    try:
        with get_connection() as con:
            # Obtener todas las solicitudes del usuario
            solicitudes = con.execute("""
                SELECT id, centro, sector, centro_costos, almacen_virtual, criticidad,
                       fecha_necesidad, justificacion, status, created_at, updated_at,
                       total_monto, aprobador_id, data_json
                FROM solicitudes
                WHERE id_usuario = ?
                ORDER BY created_at DESC
            """, (user_id,)).fetchall()

            if not solicitudes:
                return _json_error("NO_DATA", "No hay solicitudes para exportar", 404)

            # Crear workbook de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Mis Solicitudes"

            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            alignment = Alignment(horizontal="center", vertical="center")

            # Headers principales
            headers = [
                "ID", "Centro", "Sector", "Centro de Costos", "Almacén Virtual",
                "Criticidad", "Fecha Necesidad", "Justificación", "Estado",
                "Fecha Creación", "Última Actualización", "Total Estimado", "Aprobador"
            ]

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment

            # Datos de solicitudes
            current_row = 2
            for sol in solicitudes:
                try:
                    # Asegurarse de que sol sea una tupla/lista
                    if not isinstance(sol, (tuple, list)):
                        continue
                    
                    # Datos principales de la solicitud
                    ws.cell(row=current_row, column=1, value=sol[0])  # ID
                    ws.cell(row=current_row, column=2, value=sol[1] or "")  # Centro
                    ws.cell(row=current_row, column=3, value=sol[2] or "")  # Sector
                    ws.cell(row=current_row, column=4, value=sol[3] or "")  # Centro Costos
                    ws.cell(row=current_row, column=5, value=sol[4] or "")  # Almacén Virtual
                    ws.cell(row=current_row, column=6, value=sol[5] or "")  # Criticidad
                    ws.cell(row=current_row, column=7, value=sol[6] or "")  # Fecha Necesidad
                    ws.cell(row=current_row, column=8, value=sol[7] or "")  # Justificación
                    ws.cell(row=current_row, column=9, value=sol[8] or "")  # Estado
                    ws.cell(row=current_row, column=10, value=sol[9] or "")  # Fecha Creación
                    ws.cell(row=current_row, column=11, value=sol[10] or "")  # Última Actualización
                    ws.cell(row=current_row, column=12, value=sol[11] or 0)  # Total Estimado

                    # Obtener nombre del aprobador si existe
                    aprobador_name = ""
                    if len(sol) > 12 and sol[12]:  # aprobador_id
                        try:
                            aprobador = con.execute("SELECT nombre FROM usuarios WHERE id = ?", (sol[12],)).fetchone()
                            if aprobador:
                                aprobador_name = aprobador[0] or ""
                        except Exception:
                            aprobador_name = ""
                    ws.cell(row=current_row, column=13, value=aprobador_name)

                    current_row += 1

                    # Agregar items de la solicitud en filas separadas
                    data_json = sol[13] if len(sol) > 13 else None
                    if data_json and isinstance(data_json, str):
                        try:
                            data = json.loads(data_json)
                            items = data.get('items', [])
                            if items:
                                # Agregar fila de separación
                                ws.cell(row=current_row, column=1, value=f"Items de Solicitud #{sol[0]}")
                                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=13)
                                current_row += 1

                                # Headers de items
                                item_headers = ["Código", "Descripción", "Unidad", "Precio Unitario", "Cantidad", "Subtotal"]
                                for col_num, header in enumerate(item_headers, 1):
                                    cell = ws.cell(row=current_row, column=col_num, value=header)
                                    cell.font = Font(bold=True)
                                current_row += 1

                                # Datos de items
                                for item in items:
                                    if isinstance(item, dict):
                                        ws.cell(row=current_row, column=1, value=item.get('codigo', ''))
                                        ws.cell(row=current_row, column=2, value=item.get('descripcion', ''))
                                        ws.cell(row=current_row, column=3, value=item.get('unidad', ''))
                                        ws.cell(row=current_row, column=4, value=item.get('precio_unitario', 0))
                                        ws.cell(row=current_row, column=5, value=item.get('cantidad', 0))
                                        ws.cell(row=current_row, column=6, value=item.get('subtotal', 0))
                                        current_row += 1
                        except (json.JSONDecodeError, KeyError, TypeError):
                            # Si hay problemas con el JSON, continuar sin items
                            pass
                except Exception as row_error:
                    # Si hay error en una fila específica, continuar con la siguiente
                    print(f"Error procesando solicitud: {row_error}")
                    continue

            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres de ancho
                ws.column_dimensions[column_letter].width = adjusted_width

            # Guardar en buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            return send_file(
                buffer,
                as_attachment=True,
                download_name=f"mis_solicitudes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as exc:
        print(f"Error en export_solicitudes_excel: {exc}")
        import traceback
        traceback.print_exc()
        return _json_error("EXPORT_ERROR", f"Error al exportar a Excel: {exc}", 500)


@bp.get("/solicitudes/export/pdf")
def export_solicitudes_pdf():
    """Exportar todas las solicitudes del usuario autenticado a PDF"""
    user_id = _require_auth()
    if not user_id:
        return _json_error("UNAUTHORIZED", "Autenticación requerida", 401)

    try:
        with get_connection() as con:
            # Obtener todas las solicitudes del usuario
            solicitudes = con.execute("""
                SELECT id, centro, sector, centro_costos, almacen_virtual, criticidad,
                       fecha_necesidad, justificacion, status, created_at, updated_at,
                       total_monto, aprobador_id, data_json
                FROM solicitudes
                WHERE id_usuario = ?
                ORDER BY created_at DESC
            """, (user_id,)).fetchall()

            if not solicitudes:
                return _json_error("NO_DATA", "No hay solicitudes para exportar", 404)

            # Crear buffer para el PDF
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            styles = getSampleStyleSheet()

            # Estilos personalizados
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Centrado
            )

            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Heading2'],
                fontSize=14,
                spaceAfter=20,
                alignment=0  # Izquierda
            )

            normal_style = styles['Normal']

            story = []

            # Título del documento
            story.append(Paragraph("Mis Solicitudes - SPM", title_style))
            story.append(Spacer(1, 12))
            story.append(Paragraph(f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", normal_style))
            story.append(Spacer(1, 20))

            for sol in solicitudes:
                # Información de la solicitud
                story.append(Paragraph(f"Solicitud #{sol[0]}", subtitle_style))

                solicitud_data = [
                    ["Centro:", sol[1] or "-"],
                    ["Sector:", sol[2] or "-"],
                    ["Centro de Costos:", sol[3] or "-"],
                    ["Almacén Virtual:", sol[4] or "-"],
                    ["Criticidad:", sol[5] or "-"],
                    ["Fecha Necesidad:", sol[6] or "-"],
                    ["Estado:", sol[8] or "-"],
                    ["Fecha Creación:", sol[9] or "-"],
                    ["Total Estimado:", f"${sol[11] or 0:.2f}" if sol[11] else "-"],
                ]

                # Obtener nombre del aprobador
                if sol[12]:
                    aprobador = con.execute("SELECT nombre FROM usuarios WHERE id = ?", (sol[12],)).fetchone()
                    if aprobador:
                        solicitud_data.append(["Aprobador:", aprobador[0]])

                solicitud_table = Table(solicitud_data, colWidths=[100, 300])
                solicitud_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('BACKGROUND', (1, 0), (1, -1), colors.white),
                ]))
                story.append(solicitud_table)
                story.append(Spacer(1, 12))

                # Justificación
                if sol[7]:
                    story.append(Paragraph("Justificación:", styles['Heading3']))
                    story.append(Paragraph(sol[7], normal_style))
                    story.append(Spacer(1, 12))

                # Items de la solicitud
                data_json = sol[13]
                if data_json and isinstance(data_json, str):
                    try:
                        data = json.loads(data_json)
                        items = data.get('items', [])
                        if items:
                            story.append(Paragraph("Items Solicitados:", styles['Heading3']))

                            item_data = [["Código", "Descripción", "Unidad", "Precio Unit.", "Cantidad", "Subtotal"]]
                            for item in items:
                                item_data.append([
                                    item.get('codigo', ''),
                                    item.get('descripcion', ''),
                                    item.get('unidad', ''),
                                    f"${item.get('precio_unitario', 0):.2f}",
                                    str(item.get('cantidad', 0)),
                                    f"${item.get('subtotal', 0):.2f}"
                                ])

                            item_table = Table(item_data, colWidths=[60, 150, 50, 70, 60, 70])
                            item_table.setStyle(TableStyle([
                                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                                ('FONTSIZE', (0, 0), (-1, 0), 10),
                                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                                ('ALIGN', (3, 1), (5, -1), 'RIGHT'),
                                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                                ('FONTSIZE', (0, 1), (-1, -1), 9),
                                ('GRID', (0, 0), (-1, -1), 1, colors.black)
                            ]))
                            story.append(item_table)
                            story.append(Spacer(1, 20))
                    except json.JSONDecodeError:
                        pass

                # Separador entre solicitudes
                story.append(Paragraph("-" * 80, normal_style))
                story.append(Spacer(1, 20))

            # Generar PDF
            doc.build(story)
            buffer.seek(0)

            return send_file(
                buffer,
                as_attachment=True,
                download_name=f"mis_solicitudes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                mimetype="application/pdf"
            )

    except Exception as exc:
        return _json_error("EXPORT_ERROR", f"Error al exportar a PDF: {exc}", 500)






